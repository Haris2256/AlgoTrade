from openpyxl import load_workbook
from datetime import datetime
import tkinter.messagebox as mb


import variables

owned_sheet = variables.owned_sheet
transaction_sheet = variables.transaction_sheet
owned_header = variables.owned_header


def buy_stock(stock, price, amount):
    try:
        stock = stock.upper()
        price = float(price)
        amount = float(amount)
        if price < 0 or amount <= 0:
            return
        for row in range(2, owned_sheet.max_row + 1):
            if owned_sheet.cell(row=row, column=owned_header["name"]).value == stock:
                owned_sheet.cell(row=row, column=owned_header["amount"]).value += amount
                transaction_sheet.append([stock, "BUY", price, amount * price])
                owned_sheet.cell(row=row, column=owned_header["book value"]).value += price * amount
                return
        owned_sheet.append([stock, amount, amount * price])
        transaction_sheet.append([stock, "BUY", price, amount * price])
    except ValueError:
        mb.showerror("Invalid Price or Amount", f"Please enter a number. Got: {price}, {amount}")

def sell_stock(stock, price, amount):
    try:
        stock = stock.upper()
        price = float(price)
        amount = float(amount)
        if price < 0 or amount < 0:
            return
        for row in range(2, owned_sheet.max_row + 1):
            if owned_sheet.cell(row=row, column=owned_header["name"]).value == stock and owned_sheet.cell(row=row, column=2).value - amount >= 0:
                book_value = owned_sheet.cell(row=row, column=owned_header["book value"]).value
                owned_amount = owned_sheet.cell(row=row, column=owned_header["amount"]).value
                book_value -= amount * book_value / owned_amount
                owned_sheet.cell(row=row, column=owned_header["amount"]).value -= amount
                owned_sheet.cell(row=row, column=owned_header["book value"]).value = book_value
                transaction_sheet.append([stock, "SELL", price, amount * price])
                return
    except ValueError:
        mb.showerror("Invalid Price or Amount", f"Please enter a number. Got: {price}, {amount}")

def stock_value(stock, price, value, action):
    try:
        num_price = float(price)
        num_value = float(value)
        action(stock, num_price, num_value / num_price)
    except ValueError:
        mb.showerror("Invalid Price or Amount", f"Please enter a number. Got: {num_price}, {num_value}")

