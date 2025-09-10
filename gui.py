import yfinance as yf
import tkinter as tk
from openpyxl import load_workbook
from ttkwidgets.autocomplete import AutocompleteEntry
from datetime import date


import variables
import actions
import watchlist
import transactions
import details
import frame_manager

# Colours
side_colour = variables.side_colour
main_panel_colour = variables.main_panel_colour
stock_label_colour = variables.stock_label_colour
text_colour = variables.text_colour

date_format = variables.date_format

owned_sheet = variables.owned_sheet
owned_header = variables.owned_header

tickers_dict = variables.tickers_dict

stocks = load_workbook(variables.file)
STOCKS = []
owned_stocks = []

for name in tickers_dict:
    STOCKS.append(name)

# Colours
side_colour = "#363636"
main_panel_colour = "#1F1F1F"
stock_label_colour = "#2A2A2A"
text_colour = "#EEEEEE"

class Owned_Stock:
    def __init__(self, row, owned_frame):
        self.row = row
        self.name = owned_sheet.cell(row,1).value
        self.owned_frame = owned_frame
        self.ticker = tickers_dict[self.name]
        self.price = self.ticker.fast_info['last_price']
        self.prev_close = self.ticker.fast_info["previous_close"]
        self.change = (self.price - self.prev_close) / self.prev_close * 100
        self.amount = owned_sheet.cell(self.row, owned_header["amount"]).value
        self.name_lbl = tk.Label(self.owned_frame,
                                    bg=stock_label_colour,
                                    fg=text_colour,
                                    text=self.name,
                                    font=("Roboto", 16))
        self.price_lbl = tk.Label(self.owned_frame,
                                    bg=stock_label_colour,
                                    fg=text_colour,
                                    text=f"${self.price:.2f}",
                                    font=("Roboto", 16))
        self.change_lbl = tk.Label(self.owned_frame,
                                    bg=stock_label_colour,
                                    fg=text_colour,
                                    text=f"{self.change:.2f}%",
                                    font=("Roboto", 16))
        self.amount_lbl = tk.Label(self.owned_frame,
                                    bg=stock_label_colour,
                                    fg=text_colour,
                                    text=f"{self.amount:.2f}",
                                    font=("Roboto", 16))
        self.configure_frame()

    def configure_frame(self):
        for i in range(4):
            self.owned_frame.columnconfigure(i, weight=1, uniform="equal")
        self.name_lbl.grid(row=0, column=0, sticky="ew")
        self.price_lbl.grid(row=0, column=1, sticky="ew")
        self.change_lbl.grid(row=0, column=2, sticky="ew")
        self.amount_lbl.grid(row=0, column=3, sticky="ew")
        self.owned_frame.pack(padx=5,pady=(10,0), fill="x")

    def update_all(self):
        tickers_dict[self.name] = yf.Ticker(self.name)
        self.ticker = tickers_dict[self.name]
        self.price = self.ticker.fast_info["last_price"]
        self.prev_close = self.ticker.fast_info["previous_close"]
        self.change = (self.price - self.prev_close) / self.prev_close * 100
        self.amount = owned_sheet.cell(self.row, owned_header["amount"]).value
    
    def update_amount(self):
        if owned_sheet.cell(self.row, owned_header["amount"]).value is None:
            self.amount = 0
        else:
            self.amount = owned_sheet.cell(self.row, owned_header["amount"]).value
        self.amount_lbl.configure(text=f"{self.amount:.2f}")

def create_stock_frame(stock_frame, label_texts):
    for i, label_text in enumerate(label_texts):
        stock_frame.columnconfigure(i, weight=1, uniform="equal")
        label = tk.Label(
            stock_frame,
            bg=stock_label_colour,
            fg=text_colour,
            text=label_text,
            font=("Roboto", 16)
        )
        label.grid(row=0, column=i, sticky="ew")

def clean_up_sheet():
    for row in range(owned_sheet.max_row, 1, -1):
        amount = owned_sheet.cell(row, owned_header["amount"]).value
        if amount == 0 or amount is None:
            owned_sheet.delete_rows(row, 1)

def clean_up_owned():
    for owned_stock in owned_stocks:
        if owned_stock.amount == 0:
            owned_stock.owned_frame.destroy()
            owned_stocks.remove(owned_stock)

def update_amounts():
    for owned_stock in owned_stocks:
        owned_stock.update_amount()

def align_rows():
    for i, owned_stock in enumerate(owned_stocks):
        owned_stock.row = i + 2
        if owned_stock.name != owned_sheet.cell(owned_stock.row, owned_header["name"]).value:
            print(f"rows don't align: {owned_stock.name} != {owned_sheet.cell(owned_stock.row, 1).value}")

# Elements of GUI 

def initialize_gui(root):
    root.title("Stock Manager")
    root.geometry("800x600")
    root.configure(bg=main_panel_colour)

    # Side Panel
    side_panel = tk.Frame(root, bg=side_colour)
    side_panel.pack(side = "left", fill = "y")

    # Buy button
    def on_buy():
        buy_window = tk.Toplevel(root)
        buy_window.title("Buy A Stock")
        create_window(buy_window, actions.buy_stock, actions.stock_value)
        root.wait_window(buy_window)
        if len(owned_stocks) != owned_sheet.max_row - 1:
            stock_frame = tk.Frame(main_frame, bg=stock_label_colour)
            owned_stock = Owned_Stock(owned_sheet.max_row, stock_frame)
            owned_stocks.append(owned_stock)
        else:
            update_amounts()
    buy_btn = tk.Button(side_panel, text="Buy Stock", command=on_buy, font = ("Poppins", 16))
    buy_btn.pack(padx=10, pady=10)

    # Sell button
    def on_sell():
        sell_window = tk.Toplevel(root)
        sell_window.title("Sell A Stock")
        create_window(sell_window, actions.sell_stock, actions.stock_value)
        root.wait_window(sell_window)
        update_amounts()
        clean_up_sheet()
        if len(owned_stocks) != owned_sheet.max_row - 1:
            clean_up_owned()
            align_rows()
    sell_btn = tk.Button(side_panel, text="Sell Stock", command=on_sell, font = ("Poppins", 16))
    sell_btn.pack(padx=10)

    # All the frames are below
    
    manager = frame_manager.Frame_Manager()

    # Watchlist Button Command
    watchlist_frame = tk.Frame(root, bg=main_panel_colour)
    manager.add_frame("watchlist", watchlist_frame)
    def on_watchlist():
        watchlist.initialize_watchlist(watchlist_frame, main_frame)
        manager.show_frame("watchlist")
    # Watchlist Button
    watch_btn = tk.Button(side_panel, text="Watchlist", command=on_watchlist, font = ("Poppins", 16))
    watch_btn.pack(padx=10, pady=10)

    # Suggestions button
    suggestion_btn = tk.Button(side_panel, text="Suggestions",font=("Poppins", 16))
    suggestion_btn.pack(padx=10)

    # Transaction button
    def on_transaction():
        if "transaction" in manager.frames:
            manager.frames["transaction"].destroy()
        transaction_frame = tk.Frame(root, bg=main_panel_colour)
        manager.add_frame("transaction", transaction_frame)
        transactions.initialize_transactions(transaction_frame, main_frame)
        manager.show_frame("transaction")
    transaction_btn = tk.Button(
        side_panel, 
        text="Transactions", 
        command=on_transaction, 
        font=("Poppins", 16)
    )
    transaction_btn.pack(padx=10, pady=10)

    # Details Button
    details_frame = tk.Frame(root, bg=main_panel_colour)
    manager.add_frame("details", details_frame)
    def on_details():
        if "details" in manager.frames:
            manager.frames["details"].destroy()
        details_frame = tk.Frame(root, bg=main_panel_colour)
        manager.add_frame("details", details_frame)
        details.initialize_details(details_frame, main_frame)
        manager.show_frame("details")
    details_btn = tk.Button(
        side_panel, 
        text="More Details", 
        command=on_details, 
        font=("Poppins", 16)
    )
    details_btn.pack(padx=10)

    ###################################################################

    # Main panel
    main_frame = tk.Frame(root, bg=main_panel_colour)
    main_frame.pack(side = "right", fill = "both", expand = True)
    manager.add_frame("main", main_frame)
    
    # Title Frame
    title_frame = tk.Frame(main_frame, bg=stock_label_colour)
    create_stock_frame(title_frame, ["STOCK", "PRICE", "CHANGE", "AMOUNT"])
    title_frame.pack(pady=15,padx=5, fill="x")

    for row in range(2, owned_sheet.max_row + 1):
        stock_frame = tk.Frame(main_frame, bg=main_panel_colour)
        owned_stock = Owned_Stock(row,stock_frame)
        owned_stocks.append(owned_stock)

def create_window(window, action, action_value):
    window.geometry("400x400")
    window.configure(bg=main_panel_colour)
    top_frame = tk.Frame(window, bg=main_panel_colour)
    # Date Label
    label = tk.Label(top_frame, text="Date", bg=stock_label_colour, fg=text_colour, font = ("Poppins", 12))
    label.pack(pady=(10,0),fill="x", expand=True)
    # Date Entry
    date_entry = tk.Entry(top_frame, font=("Poppins", 8))
    today_str = date.today().strftime(date_format)
    date_entry.insert(0, today_str)
    date_entry.pack(pady=10)
    # Stock Name Label
    label = tk.Label(top_frame, text="Stock Name", bg=stock_label_colour, fg=text_colour, font = ("Poppins", 12))
    label.pack(pady=(10,0),fill="x", expand=True)
    # Stock Name Entry
    name_entry = AutocompleteEntry(top_frame, completevalues=STOCKS)
    name_entry.pack(pady=10)
    # Stock Price Label
    label = tk.Label(top_frame, text="Stock Price", bg=stock_label_colour, fg=text_colour, font = ("Poppins", 12))
    label.pack(pady=(10,0),fill="x",expand=True)
    # Stock Price Entry
    price_entry = tk.Entry(top_frame, font=("Poppins", 8))
    price_entry.pack(pady=(10,0))
    # Stock Amount Label
    label = tk.Label(top_frame, text="Amount", bg=stock_label_colour, fg=text_colour, font = ("Poppins", 12))
    label.pack(pady=(10,0), fill="x", expand=True)
    # Pack top frame
    top_frame.pack(padx=10,fill="x")
    # Bottom Frame
    bottom_frame = tk.Frame(window, bg=main_panel_colour)
    for i in range(0,2):
        bottom_frame.grid_columnconfigure(i, weight=1)
    # Share Amount Entry
    amount_entry = tk.Entry(bottom_frame, font=("Poppins", 8))
    amount_entry.grid(row=0,column=0, padx=10, sticky="ew")
    # Share Amount Button
    def on_action_amount():
        action(date_entry.get(), name_entry.get(), price_entry.get(), amount_entry.get())
        window.destroy()
    stock_amount = tk.Button(bottom_frame, text="Share Amount",
                             command=on_action_amount,
                             font = ("Poppins", 12))
    stock_amount.grid(row=1, column=0, pady=5, padx=10, sticky="ew")
    # Total Value Entry
    value_entry = tk.Entry(bottom_frame, font=("Poppins", 8))
    value_entry.grid(row=0, column=1, padx=10,sticky="ew")
    # Total Value Button
    def on_action_value():
        action_value(date_entry.get(), name_entry.get(), price_entry.get(), value_entry.get(), action)
        window.destroy()
    stock_value = tk.Button(bottom_frame, text="Total Value",
                            command=on_action_value,
                            font = ("Poppins", 12))
    stock_value.grid(row=1, column=1, pady=5, padx=10, sticky="ew")
    # Pack bottom frame
    bottom_frame.pack(padx=10, pady=10,fill="y",expand=True)

