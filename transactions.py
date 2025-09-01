import tkinter as tk
from openpyxl import load_workbook
from datetime import datetime
import tkinter.messagebox as mb

import variables

transaction_sheet = variables.transaction_sheet
transaction_header = variables.transaction_header

main_panel_colour = variables.main_panel_colour
stock_label_colour = variables.stock_label_colour
text_colour = variables.text_colour

transactions = []


class Transaction:
    def __init__(self, row, frame):
        self.row = row
        self.frame = frame
        self.frame.configure(bg=stock_label_colour, bd=1, relief="raised")
        for i in range(transaction_sheet.max_column):
            frame.columnconfigure(i, weight=1,uniform="equal")
        
        frame.pack(padx=5, pady=2, fill="x")
        # Values from the sheet
        self.date = transaction_sheet.cell(row, transaction_header["date"]).value
        self.name = transaction_sheet.cell(row, transaction_header["name"]).value
        self.action = transaction_sheet.cell(row, transaction_header["action"]).value
        self.price = transaction_sheet.cell(row, transaction_header["price"]).value
        self.value = transaction_sheet.cell(row, transaction_header["value"]).value

        # Labels
        self.name_lbl = tk.Label(frame, text=self.name, bg=stock_label_colour, fg=text_colour, font=("Roboto", 12))
        self.action_lbl = tk.Label(frame, text=self.action, bg=stock_label_colour, fg=text_colour, font=("Roboto", 12))
        self.price_lbl = tk.Label(frame, text=f"${self.price:.2f}", bg=stock_label_colour, fg=text_colour, font=("Roboto", 12))
        self.value_lbl = tk.Label(frame, text=f"${self.value:.2f}", bg=stock_label_colour, fg=text_colour, font=("Roboto", 12))
        self.date_lbl = tk.Label(frame, text=self.date.strftime("%Y-%m-%d"), bg=stock_label_colour, fg=text_colour, font=("Roboto", 12))

        # Pack Labels
        self.date_lbl.grid(row=0, column=0)
        self.name_lbl.grid(row=0, column=1)
        self.action_lbl.grid(row=0, column=2)
        self.price_lbl.grid(row=0, column=3)
        self.value_lbl.grid(row=0, column=4)

        # Entries for edit mode
        self.date_entry = tk.Entry(frame, bg=stock_label_colour, fg=text_colour,font=("Roboto", 12))
        self.name_entry = tk.Entry(frame, bg=stock_label_colour, fg=text_colour,font=("Roboto", 12))
        self.action_entry = tk.Entry(frame, bg=stock_label_colour, fg=text_colour, font=("Roboto", 12))
        self.price_entry = tk.Entry(frame, bg=stock_label_colour, fg=text_colour, font=("Roboto", 12))
        self.value_entry = tk.Entry(frame, bg=stock_label_colour, fg=text_colour, font=("Roboto", 12))

        # Pre-fill entries with current values
        self.date_entry.insert(0, self.date.strftime("%Y-%m-%d"))
        self.name_entry.insert(0, self.name)
        self.action_entry.insert(0, self.action)
        self.price_entry.insert(0, self.price)
        self.value_entry.insert(0, self.value)

        # Delete Button
        self.del_btn = tk.Button(frame, bg=stock_label_colour, fg=text_colour, text="X", command=self.delete)
    
    def convert_edit(self):
        self.name_lbl.grid_forget()
        self.action_lbl.grid_forget()
        self.price_lbl.grid_forget()
        self.value_lbl.grid_forget()
        self.date_lbl.grid_forget()

        self.date_entry.grid(row=0, column=0)
        self.name_entry.grid(row=0, column=1)
        self.action_entry.grid(row=0, column=2)
        self.price_entry.grid(row=0, column=3)
        self.value_entry.grid(row=0, column=4)

        self.del_btn.grid(row=0, column=5)

    def save(self):
        new_name = self.name_entry.get()
        new_action = self.action_entry.get().upper()
        if new_action != "BUY" and new_action != "SELL":
            mb.showerror("Invalid Action", f"Please enter BUY or SELL. Got: {new_action}")
            new_action = self.action
        try:
            new_date = self.date_entry.get()
            new_date = datetime.strptime(new_date, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            mb.showerror("Invalid Date", f"Please enter date as YYYY-MM-DD. Got: {new_date}")
            new_date = self.date
        try:
            new_price = self.price_entry.get()
            new_price = float(new_price)
        except (ValueError, TypeError):
            mb.showerror("Invalid Price", f"Please enter a number for price. Got: {new_price}")
            new_price = self.price
        try:
            new_value = self.value_entry.get()
            new_value  = float(new_value)
        except (ValueError, TypeError):
            mb.showerror("Invalid Value", f"Please enter a number for value. Got: {new_value}")
            new_value = self.value

        transaction_sheet.cell(self.row, transaction_header["date"]).value = new_date
        transaction_sheet.cell(self.row, transaction_header["name"]).value = new_name
        transaction_sheet.cell(self.row, transaction_header["action"]).value = new_action
        transaction_sheet.cell(self.row, transaction_header["price"]).value = new_price
        transaction_sheet.cell(self.row, transaction_header["value"]).value = new_value

        self.date, self.name, self.action, self.price, self.value = new_date, new_name, new_action, new_price, new_value

        self.date_lbl.config(text=new_date.strftime("%Y-%m-%d"))
        self.name_lbl.config(text=new_name)
        self.action_lbl.config(text=new_action)
        self.price_lbl.config(text=f"${new_price:.2f}")
        self.value_lbl.config(text=f"${new_value:.2f}")

        self.date_entry.grid_forget()
        self.name_entry.grid_forget()
        self.action_entry.grid_forget()
        self.price_entry.grid_forget()
        self.value_entry.grid_forget()

        self.del_btn.grid_forget()

        self.date_lbl.grid(row=0, column=0)
        self.name_lbl.grid(row=0, column=1)
        self.action_lbl.grid(row=0, column=2)
        self.price_lbl.grid(row=0, column=3)
        self.value_lbl.grid(row=0, column=4)
    
    def delete(self):
        transactions.remove(self)
        transaction_sheet.delete_rows(self.row, 1)
        self.frame.destroy()



def initialize_transaction_window(window):
    window.geometry("700x500")

    button_frame = tk.Frame(window, bg=main_panel_colour)
    button_frame.pack(fill="x")

    # Exit Button
    def on_exit():
        transactions.clear()
        window.destroy()

    exit_btn = tk.Button(button_frame, text="Exit",command=on_exit, font=("Poppins", 12),width=10)
    exit_btn.pack(side="left", anchor="n",pady=10, padx=10)

    def on_edit():
        for transaction in transactions:
            transaction.convert_edit()
    edit_btn = tk.Button(button_frame, text="Edit", command=on_edit, font=("Poppins", 12),width=10)
    edit_btn.pack(side="right", anchor="n",pady=10, padx=10)

    def on_save():
        for transaction in transactions:
            transaction.save()
    save_btn = tk.Button(button_frame, text="Save", command=on_save, font=("Poppins", 12),width=10)
    save_btn.pack(side="right", anchor="n",pady=10, padx=10)

    container = tk.Frame(window)         # holds both canvas and scrollbar
    container.pack(fill="both", expand=True)

    canvas = tk.Canvas(container, bg=main_panel_colour, highlightthickness=0)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    # link canvas and scrollbar
    canvas.configure(yscrollcommand=scrollbar.set)

    # --- frame inside canvas where content goes ---
    transaction_frame = tk.Frame(canvas, bg=main_panel_colour)

    # put that frame into the canvas
    window_id = canvas.create_window((0,0), window=transaction_frame, anchor="nw")

    # update scrollregion when widgets inside transaction_frame change size
    def on_frame_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    transaction_frame.bind("<Configure>", on_frame_configure)

    def on_canvas_configure(event):
        canvas.itemconfig(window_id, width=event.width)  # stretch to canvas width

    canvas.bind("<Configure>", on_canvas_configure)

    # Enable mousewheel scrolling
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    canvas.bind_all("<MouseWheel>", _on_mousewheel)   # Windows/Mac
    canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))  # Linux
    canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))   # Linux

    header_frame = tk.Frame(transaction_frame, bg=stock_label_colour, bd=1, relief="raised")
    header_frame.pack(fill="x", padx=5, pady=10)
    for i in range(1,transaction_sheet.max_column + 1):
        header_frame.columnconfigure(i, weight=1, uniform="equal")
        name = transaction_sheet.cell(1, i).value
        name = name.upper()
        label = tk.Label(header_frame, bg=stock_label_colour, fg=text_colour,
                         text=name, font=("Roboto", 12))
        label.grid(row=0,column=i, sticky="ew")

    # Transction rows
    for row in range(2, transaction_sheet.max_row + 1):
        frame = tk.Frame(transaction_frame)
        transaction = Transaction(row, frame)
        transactions.append(transaction)


