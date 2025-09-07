import tkinter as tk
from openpyxl import load_workbook
from datetime import datetime
import tkinter.messagebox as mb
from tkinter import ttk

import scrollarea
import variables

transaction_sheet = variables.transaction_sheet
transaction_header = variables.transaction_header

owned_sheet = variables.owned_sheet
owned_header = variables.owned_header

main_panel_colour = variables.main_panel_colour
stock_label_colour = variables.stock_label_colour
text_colour = variables.text_colour

date_format = variables.date_format

stocks_dict = {}

class Stock_Detail:
    def __init__(self, name):
        self.name = name
        self.amount = 0
        self.book_value = 0
        self.realized_return = 0
        self.transactions = []

    def add_transaction(self, row):
        date = transaction_sheet.cell(row, transaction_header["date"]).value
        name = transaction_sheet.cell(row, transaction_header["name"]).value
        action = transaction_sheet.cell(row, transaction_header["action"]).value
        price = transaction_sheet.cell(row, transaction_header["price"]).value
        value = transaction_sheet.cell(row, transaction_header["value"]).value
        self.transactions.append((date, name, action, price, value))
        if action == "BUY":
            self.amount += value / price
            self.book_value += value
        else:
            average_price = self.book_value / self.amount
            amount = value / price
            self.realized_return += (price - average_price) * amount
            self.book_value -= average_price * amount
            self.amount -= amount

def calculate_realized_return():
    sum = 0
    for detail in stocks_dict.values():
        sum += detail.realized_return
    return round(sum,2)

def initialize_details(details_frame, main_frame):
    details_frame.pack(side="right", fill="both",expand=True)

    content_frame = scrollarea.make_scrollarea(details_frame, main_panel_colour)
    content_frame.pack(fill="both",expand=True)

    # Create Stock Objects:
    for row in range(2, transaction_sheet.max_row + 1):
        stock_name = transaction_sheet.cell(row, transaction_header["name"]).value

        # Ensure one Stock_Detail per name
        if stock_name not in stocks_dict:
            stocks_dict[stock_name] = Stock_Detail(stock_name)

        # Add txn into that stock’s detail object
        stocks_dict[stock_name].add_transaction(row)

    # Exit button
    def on_exit():
        details_frame.destroy()
        stocks_dict.clear()
        main_frame.pack(side="right", fill="both", expand=True)

    exit_btn = tk.Button(
        content_frame, 
        text="Exit",
        command=on_exit, 
        font=("Poppins", 12),
        width=10
    )
    exit_btn.pack(anchor="w", padx=10, pady=10)

    # Overview
    overview_frame = tk.Frame(content_frame, bg=main_panel_colour)
    overview_frame.pack(fill="x", expand=True)

    # Total Return
    

    # Realized Return
    total_realized_return = calculate_realized_return()
    label = tk.Label(
        overview_frame, 
        bg=stock_label_colour,
        fg=text_colour,
        text=f"Realized Return: ${total_realized_return}"
    )
    label.pack(anchor="w")






