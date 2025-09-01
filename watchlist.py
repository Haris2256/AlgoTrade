import yfinance as yf
import tkinter as tk
from openpyxl import load_workbook

import variables

# Colours
side_colour = variables.side_colour
main_panel_colour = variables.main_panel_colour
stock_label_colour = variables.stock_label_colour
text_colour = variables.text_colour

watchlist_sheet = variables.watchlist_sheet

num_stocks_per_column = variables.num_stocks_per_column
num_stocks_per_row = variables.num_stocks_per_row

watchlist_stocks = []

# --- Watchlist ---

class Watchlist_Stock:
    def __init__(self, name, stock_frame, row):
        self.name = name
        self.row = row
        self.ticker = yf.Ticker(name)
        self.current_price = self.ticker.fast_info["last_price"]
        self.prev_close = self.ticker.fast_info["previous_close"]
        self.pct_change = (self.current_price - self.prev_close) / self.prev_close * 100
        self.hist = self.ticker.history(period="5y")
        self.changes = self.pct_changes_multi()
        self.stock_frame = stock_frame
        stock_frame.configure(bg=stock_label_colour, bd=1, relief="raised")
        self.price_frame = tk.Frame(self.stock_frame, bg=stock_label_colour)
        self.name_lbl  = tk.Label(self.stock_frame, text=self.name,
                            bg=stock_label_colour, fg=text_colour)
        self.price_lbl = tk.Label(self.price_frame, text=f"${self.current_price:.2f}", 
                            bg=stock_label_colour, 
                            fg="lightgreen" if self.pct_change >= 0 else "red")
        self.pct_change_lbl = tk.Label(self.price_frame, text=f"{self.pct_change:.2f}%", 
                                    bg=stock_label_colour, 
                                    fg="lightgreen" if self.pct_change >= 0 else "red")
        self.week_lbl  = tk.Label(self.stock_frame, 
                            text=f"1W: {self.changes['1W']:.2f}%",
                            bg=stock_label_colour,
                            fg="lightgreen" if self.changes['1W'] >= 0 else "red")

        self.month_lbl = tk.Label(self.stock_frame,
                            text=f"1M: {self.changes['1M']:.2f}%",
                            bg=stock_label_colour,
                            fg="lightgreen" if self.changes['1M'] >= 0 else "red")

        self.year_lbl  = tk.Label(self.stock_frame,
                            text=f"1Y: {self.changes['1Y']:.2f}%",
                            bg=stock_label_colour,
                            fg="lightgreen" if self.changes['1Y'] >= 0 else "red")

        self.fivey_lbl = tk.Label(self.stock_frame,
                            text=f"5Y: {self.changes['5Y']:.2f}%",
                            bg=stock_label_colour,
                            fg="lightgreen" if self.changes['5Y'] >= 0 else "red")
        self.configure_frame()
    
    def pct_changes_multi(self):
        if self.hist.empty:
            return {"1W": 0, "1M": 0, "1Y": 0, "5Y": 0}
        close_prices = self.hist["Close"]
        def pct_change(start_price, end_price):
            return (end_price - start_price) / start_price * 100
        last_close = close_prices.iloc[-1]  # most recent closing price
        # Now compute % changes:
        changes = {}
        # 1 Week (~5 trading days ago)
        if len(close_prices) >= 5:
            changes["1W"] = pct_change(close_prices.iloc[-5], last_close)
        else:
            changes["1W"] = 0
        # 1 Month (~21 trading days)
        if len(close_prices) >= 21:
            changes["1M"] = pct_change(close_prices.iloc[-21], last_close)
        else:
            changes["1M"] = 0
        # 1 Year (~252 trading days)
        if len(close_prices) >= 252:
            changes["1Y"] = pct_change(close_prices.iloc[-252], last_close)
        else:
            changes["1Y"] = 0
        # 5 Years (first vs last in the DataFrame)
        changes["5Y"] = pct_change(close_prices.iloc[0], last_close)
        return changes

    def configure_frame(self):
        self.price_lbl.pack(side="left")
        self.pct_change_lbl.pack(side="right")

        # pack labels vertically
        self.name_lbl.pack(pady=2)
        self.price_frame.pack(padx=5, fill="x", expand=True)
        self.week_lbl.pack(anchor="e", padx=5)
        self.month_lbl.pack(anchor="e", padx=5)
        self.year_lbl.pack(anchor="e", padx=5)
        self.fivey_lbl.pack(anchor="e", padx=5, pady=(0,5))
    def update(self):
        # recompute percent changes
        self.ticker = yf.Ticker(self.name)
        self.hist = self.ticker.history(period="5y")
        self.current_price = self.ticker.fast_info["last_price"]
        self.prev_close = self.ticker.fast_info["previous_close"]
        self.pct_change = (self.current_price - self.prev_close) / self.prev_close * 100
        self.changes = self.pct_changes_multi()

        # update each label
        self.price_lbl.config(
            text=f"${self.current_price:.2f}", 
            fg="lightgreen" if self.pct_change >= 0 else "red"
        )
        self.pct_change_lbl.config(
            text=f"{self.pct_change:.2f}%", 
            fg="lightgreen" if self.pct_change >= 0 else "red"
        )
        self.week_lbl.config(
            text=f"1W: {self.changes['1W']:.2f}%",
            fg="lightgreen" if self.changes['1W'] >= 0 else "red"
        )
        self.month_lbl.config(
            text=f"1M: {self.changes['1M']:.2f}%",
            fg="lightgreen" if self.changes['1M'] >= 0 else "red"
        )
        self.year_lbl.config(
            text=f"1Y: {self.changes['1Y']:.2f}%",
            fg="lightgreen" if self.changes['1Y'] >= 0 else "red"
        )
        self.fivey_lbl.config(
            text=f"5Y: {self.changes['5Y']:.2f}%",
            fg="lightgreen" if self.changes['5Y'] >= 0 else "red"
        )
    def regrid(self):
        self.stock_frame.grid(row=(self.row-1)//num_stocks_per_row, 
                         column=(self.row-1)%num_stocks_per_row, 
                         sticky="ew", padx=5)

def align_rows():
    for i, watchlist_stock in enumerate(watchlist_stocks):
        watchlist_stock.row = i + 1

def clean_up_sheet():
    for row in range(watchlist_sheet.max_row, 1, -1):
        if watchlist_sheet.cell(row, 1) is None:
            watchlist_sheet.delete_rows(row, 1)

def grid_watchlist_stocks():
    for watchlist_stock in watchlist_stocks:
        watchlist_stock.regrid()

def initialize_watchlist(watchlist_frame, main_frame):
    if watchlist_stocks:
        watchlist_frame.pack(side="right", fill="both", expand=True)
        return
    def add_to_watchlist(stock_name, row):
        stock_frame = tk.Frame(watch_frame)
        watchlist_stock = Watchlist_Stock(stock_name, stock_frame, row)
        watchlist_stock.regrid()
        watchlist_stocks.append(watchlist_stock)

    # Inner Frames
    button_frame = tk.Frame(watchlist_frame, bg=main_panel_colour)
    watch_frame = tk.Frame(watchlist_frame, bg=main_panel_colour)

    # Pack Inner Frames
    button_frame.pack(fill="x")
    watch_frame.pack(fill="both", expand=True)

    # Pack Main Frame
    watchlist_frame.pack(side="right", fill="both", expand=True)

    # Exit Button
    def on_exit():
        watchlist_frame.pack_forget()
        main_frame.pack(side = "right", fill = "both", expand = True)

    exit_btn = tk.Button(button_frame, text="Exit",command=on_exit, font=("Poppins", 12),width=6)
    exit_btn.pack(side="left", anchor="n",pady=10, padx=10)

    # Add Button
    def on_add():
        new_stock = new_stock_entry.get()
        new_stock = new_stock.upper()
        for row in range(1,watchlist_sheet.max_row + 1):
            if new_stock == watchlist_sheet.cell(row, 1).value:
                return
        try:
            ticker = yf.Ticker(new_stock)
            info = ticker.fast_info
        except Exception:
            return
        watchlist_sheet.cell(watchlist_sheet.max_row + 1, 1).value = new_stock
        add_to_watchlist(new_stock, watchlist_sheet.max_row)

    add_frame = tk.Frame(button_frame, bg=main_panel_colour)
    add_to_watchlist_btn = tk.Button(add_frame, text="Add to Watchlist", 
                                     command=on_add, font=("Poppins", 12))
    add_to_watchlist_btn.pack(pady=10, padx=10)
    new_stock_entry = tk.Entry(add_frame, font=("Roboto", 8))
    new_stock_entry.pack()
    add_frame.pack(side="right")

     # Remove Button
    def on_remove():
        remove_stock = remove_stock_entry.get()
        remove_stock = remove_stock.upper()
        for watchlist_stock in watchlist_stocks:
            if watchlist_stock.name == remove_stock:
                watchlist_stock.stock_frame.destroy()
                watchlist_sheet.delete_rows(watchlist_stock.row,1)
                watchlist_stocks.remove(watchlist_stock)
        align_rows()
        grid_watchlist_stocks()
    remove_frame = tk.Frame(button_frame, bg=main_panel_colour)
    remove_from_watchlist_btn = tk.Button(remove_frame, text="Remove from Watchlist", 
                                     command=on_remove, font=("Poppins", 12))
    remove_from_watchlist_btn.pack(pady=10, padx=10)
    remove_stock_entry = tk.Entry(remove_frame, font=("Roboto", 8))
    remove_stock_entry.pack()
    remove_frame.pack(side="right")

    # Update Button
    def update_watchlist_stocks():
        for watchlist_stock in watchlist_stocks:
            watchlist_stock.update()
    update_btn = tk.Button(button_frame, text="Update", command=update_watchlist_stocks,
                           font=("Poppins", 12))
    update_btn.pack(side="right", anchor="n",padx=10, pady=10)

    # Configure Main Frame 
    for i in range(num_stocks_per_column):
        watch_frame.grid_columnconfigure(i, weight=1)
    # Configure correct number of rows based on number of items in watchlist
    for i in range(num_stocks_per_row):
        watch_frame.grid_rowconfigure(i, weight=1)

    for row in range(1, watchlist_sheet.max_row + 1):
        add_to_watchlist(watchlist_sheet.cell(row,1).value, row)
    
    