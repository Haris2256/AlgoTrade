from openpyxl import load_workbook
import yfinance as yf

file = "stocks.xlsx"
stocks = load_workbook(file)

# Spreadsheets
owned_sheet = stocks["owned stocks"]
watchlist_sheet = stocks["watchlist"]
transaction_sheet = stocks["transactions"]

# Sheet Headers
owned_header = {cell.value: idx for idx, cell in enumerate(owned_sheet[1], start=1)}
transaction_header = {cell.value: idx for idx, cell in enumerate(transaction_sheet[1], start=1)}

# Tickers Dictionary so yfinance isn't accessed too often
tickers_dict = {}
# owned
for row in range(2, owned_sheet.max_row + 1):
    ticker = owned_sheet.cell(row=row, column=owned_header["name"]).value
    if ticker and ticker not in tickers_dict:
        tickers_dict[ticker] = yf.Ticker(ticker)

# watchlist
for row in range(2, watchlist_sheet.max_row + 1):
    ticker = watchlist_sheet.cell(row=row, column=1).value
    if ticker and ticker not in tickers_dict:
        tickers_dict[ticker] = yf.Ticker(ticker)

# transactions
for row in range(2, transaction_sheet.max_row + 1):
    ticker = transaction_sheet.cell(row=row, column=transaction_header["name"]).value
    if ticker and ticker not in tickers_dict:
        tickers_dict[ticker] = yf.Ticker(ticker)


# Colours
side_colour = "#363636"
main_panel_colour = "#1F1F1F"
stock_label_colour = "#2A2A2A"
text_colour = "#EEEEEE"

# Watchlist Configuration
num_stocks_per_row = 5
num_stocks_per_column = 5

# Date Format
date_format = "%Y-%m-%d"