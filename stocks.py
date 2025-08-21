import pandas as pd
import yfinance as yf
import tkinter as tk
import actions
from openpyxl import load_workbook


file = "prices.xlsx"
stocks = load_workbook(file)

root = tk.Tk()
root.title("Stock Manager")
root.geometry("800x600")

# Colours
side_colour = "#363636"
main_panel_colour = "#1F1F1F"
stock_label_colour = "#2A2A2A"
text_colour = "#AAAAAA"


# Side Panel
side_panel = tk.Frame(root, bg=side_colour)
side_panel.pack(side = "left", fill = "y")

# Buy button
buybtn = tk.Button(side_panel, text="Buy Stock", font = ("Poppins", 16))
buybtn.pack(padx=10, pady=10)

# Sell button
sellbtn = tk.Button(side_panel, text="Sell Stock", font = ("Poppins", 16))
sellbtn.pack(padx=10)

# Add to watchlist
watchbtn = tk.Button(side_panel, text="Add to Watchlist", font = ("Poppins", 16))
watchbtn.pack(padx=10, pady=10)

# Suggestions button
sellbtn = tk.Button(side_panel, text="Suggestions", font = ("Poppins", 16))
sellbtn.pack(padx=10)

###################################################################

# Main panel
main_frame = tk.Frame(root, bg=main_panel_colour)
main_frame.pack(side = "right", fill = "both", expand = True)
for i in range(0,3):
    main_frame.grid_columnconfigure(i, weight=1)
label = tk.Label(main_frame, text="STOCK", bg=stock_label_colour, fg=text_colour,font=("Roboto", 16))
label.grid(row=0, column=0, sticky="ew", padx=(5,0), pady=20)
label = tk.Label(main_frame, text="PRICE", bg=stock_label_colour, fg=text_colour,font=("Roboto", 16))
label.grid(row=0, column=1, sticky="ew")
label = tk.Label(main_frame, text="CHANGE", bg=stock_label_colour, fg=text_colour,font=("Roboto", 16))
label.grid(row=0, column=2, sticky="ew")
label = tk.Label(main_frame, text="AMOUNT", bg=stock_label_colour, fg=text_colour,font=("Roboto", 16))
label.grid(row=0, column=3, sticky="ew", padx=(0,5))

sheet = stocks.active
for row in range(2, sheet.max_row + 1):
    # Stock Name
    stock_name = sheet.cell(row, 1).value
    exchange = sheet.cell(row, 2).value
    stock_name += f".{exchange}"
    label = tk.Label(main_frame, text=stock_name, bg=stock_label_colour, fg=text_colour,font=("Roboto", 16))
    label.grid(row=row, column=0, sticky="ew", padx=(5,0), pady=5)
    # Stock Price
    stock = yf.Ticker(stock_name)
    current_price = stock.fast_info['last_price']
    output = f"{stock_name} {round(current_price,2)}"
    label = tk.Label(main_frame, text=f"{round(current_price, 2)}", bg=stock_label_colour, fg=text_colour,font=("Roboto", 16))
    label.grid(row=row, column=1, sticky="ew")
    # Stock Change
    prev_close = stock.fast_info["previous_close"]
    pct_change = round(((current_price - prev_close) / prev_close) * 100, 2)
    label = tk.Label(main_frame, text=f"{pct_change}%", bg=stock_label_colour, fg=text_colour, font=("Roboto", 16))
    label.grid(row=row, column=2, sticky="ew")
    # Stock Amount
    stock_amount = sheet.cell(row, 3).value
    label = tk.Label(main_frame, text=stock_amount, bg=stock_label_colour, fg=text_colour, font=("Roboto", 16))
    label.grid(row=row, column=3, sticky="ew", pady=5, padx=(0,5))



root.mainloop()
